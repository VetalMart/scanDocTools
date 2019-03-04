import os

from openpyxl import load_workbook
from easygui import fileopenbox

import data



def splitName(s):
    """функция для деланья из полных имен-фамилий с инициалами"""
    l = s.split('\n')
    l.remove('')
    lN = []
    for i in l:
        j = i.split(' ')
        if len(j) == 2:
            lN.append('{0}{1}'.format(j[1], j[0]))
        else:
            lN.append('{0}.{1}.{2}'.format(j[1][0], j[2][0], j[0]))
    return lN

"""
Код для фильтрации текста из бухгалтеркого файла, который сбрасивает
Левикина.
"""
def obj_with_vrez(s) -> dict:
    """Виделяет с текста объекты в которых выполнена врезка.

        И стандартные и не стандартные.
        s - сырой текст с екселя;
    """
    raw_list = s.split('03217')  # делим на объекты по коду, для всех один 
    # делим список на три части: номер, адрес, остальная инфа
    temp_list = [i.split('\n', maxsplit=2) for i in raw_list]
    vrez_dict = {}      # словарь с объектами где есть врезка 
    # отбираем объекты где есть врезка
    for i in temp_list:
        for j in i:
            # 60500ID - код врезки
            # 60600ID - код тех.надзора
            # с врезкой, без тех.надзора
            if ('60500ID' in j 
                and '60600ID' not in j):
                # ключ - номер, значение - адрес; 
                # обрезаем пробелы
                vrez_dict.setdefault(i[0].strip(), i[1].strip())

    return vrez_dict

def st_obj(d, standart=0) -> dict:
    """Отбирает объекты внешнего присоединения.

    d - словарь со всеми присоединениями; 
    standart - флаг: если 1 - стандартное; 
                     если 0 - нестандартное(по умолчанию);
    """
    if standart==1:
        check_type = '250133'   # код стандартного присоединения
    elif standart==0:
        check_type = '250233'   # код нестандартного присоединения
    
    temp_dict = {}              # промежуточный словарь
    
    for k, v in d.items():
        if check_type in k:
            temp_dict.setdefault(k, v)

    return temp_dict

def whats_in_folder(d, needs=0) -> dict:
    """Возвращает необходимое. 

    Функция шерстит папку Кравченко, какие там акты по нестандартним есть,
    и в зависимости от этого, может потребоваться следующая инфа:
    1. Какие акты есть;
    2. Каких актов нет;

    Во избежании лишнего гемора с переходами по файловой системе, 
    мы подразумеваем что уже находимся в папке с екселевскими файлами, 
    по нестандартному присоединению, которые делала Кравченко;

    d - словарь с объектами с нестандартным присоединением;
    needs - флаг: 0 - поиск в папке, каких актов нет;
                  1 - поиск в папке, какие акты есть;  
    """
    # удаляем расширения, подразумеваем что они еще .xls
    # и создаем сразу список
    set_folder_files = {i.rstrip('.xls') for i in os.listdir()}
    temp_dict = {}
    diff = set(d).difference(set_folder_files)
    inter = set(d).intersection(set_folder_files)
    if needs==0:
        for i in diff:
            temp_dict.setdefault(i, d[i])            
    elif needs==1:
        for i in inter:
            temp_dict.setdefault(i, d[i])             

    return temp_dict

def dict_to_txt_file(d)->None:
    """Печатает в текстовый файл словарь
    
    - печатает с кодировкой cp1251 для windows;
    - нужно будет ввести имя файла.
    - в файле информация идет так: "Ключ" {табуляция} "значение" - для 
        удобства вставки в ексель;

    d - словарь который нужно завернуть в файл"""
    name = input("Enter file name(no extention)> ")
    with open('{0}.txt'.format(name), 'w', encoding='cp1251') as f:
        for k, v in d.items():
            f.write('{0}\t{1}\n'.format(k, v))

def list_to_txt_file(l)->None:
    """Печатает в текстовый файл список
    
    - печатает с кодировкой cp1251 для windows;
    - нужно будет ввести имя файла.
    - в файле информация идет так: "Ключ" {табуляция} "значение" - для 
        удобства вставки в ексель;

    l - список который нужно завернуть в файл"""
    name = input("Enter file name(no extention)> ")
    with open('{0}.txt'.format(name), 'w', encoding='cp1251') as f:
        for k in l:
            f.write('{0}\n'.format(k))

def make_list_from_excel_str(s)->list:
    """Делает список из копии текста с екселя"""
    tmp_str = s.split('\n')
    if tmp_str[-1] == '':
        tmp_str.pop(-1)
    return tmp_str

def whitespace_remover(s):
    """Delete whitespace if it repeate > 2 times, or after \n"""
    tmp_s = ''
    c = 0
    for i in s:
        if not i.isspace():
            tmp_s += i
            c = 0
        elif i.isspace():
            if c == 0:
                tmp_s += i
                c = 1
            else:
                pass

    return tmp_s.lower()

def temp_remover(s, code)->list:
    vas_tmp = whitespace_remover(s).replace('\n', '').split(code)
    vas = [i.strip().split(' ') for i in vas_tmp]
    for i in vas:
        if i == ['']:
            vas.remove(i)
    return vas

def list_from_excel_file(num, left_up_bound, right_down_bound)->list:
    """Возвращает список значений из выбраного файла нужного диапазона.
    """
    wb = load_workbook(fileopenbox())   #choose file
    #for i in enumerate(wb.sheetnames):  #check worksheets
     #   print(i)
    #num = int(input("Type number of sheets you want.> "))
    ws = wb[wb.sheetnames[num]]         #choose require worksheets
    #choose bounds
    #left_up_bound = str(input("Type left-up cell bound> "))
    #right_down_bound = str(input("Type right down cell bound> "))
    cell_range = ws[left_up_bound:right_down_bound]
    cell_grid = []
    #go through rows
    for row in cell_range:
        tmp = []
        for cell in row:
            tmp.append(cell.value)
        cell_grid.append(tmp)

    return cell_grid
    
if __name__ == "__main__":
    #vas, fas, vas_fas, pysanko = [], [], [], []
    #vas = list_from_excel_file(2, 'A2', 'H343')
    #fas = list_from_excel_file(3, 'A2', 'H140')
    #vas_fas = list_from_excel_file(1, 'A2', 'H378')
    ksv = list_from_excel_file(1, 'A3', 'H279')
    pysanko = list_from_excel_file(0, 'E5', 'H2007')
    #tmp_list = [vas, fas, vas_fas, pysanko]

    #total = []
    #for i in tmp_list[:3]:
     #   total.extend(i)
    
    match = []
    for i in ksv:#total:
        for j in pysanko:
            if i[3] == None:
                print('!!!!!!!!!!!!')
                print(i)
                print('!!!!!!!!!!!!')
            elif (i[0] in j[1] and
                    i[3].split(' ')[0] in j[0] and
                    i[7] == j[3]):
                print(j)
                #match.append(j)
    """
    list_to_txt_file(match)
    
    vas = temp_remover(data.vas, '06$')
    fas = temp_remover(data.fas, '23$')
    vas_fas = [i.strip().split(' ') for i in data.vas_fas.splitlines()]
    l = [vas, fas, vas_fas]
    total = []
    for i in l:
        total.extend(i)     #my dog

    req_list = []
    numb_dog = []
    for i in total:
        for j in data.pysanko.lower().split('\n'):
            if (#i[0] in j and 
                i[1] in j):# and 
                #i[-1] in j):
                req_list.append(j)
                numb_dog.append(i)
                #print(i[0], i[1])
                print(j)
    #list_to_txt_file(req)
    """